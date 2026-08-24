from __future__ import annotations

import importlib.util
import json
from pathlib import Path
import sys
import tempfile
import unittest


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "schedule_night_shifts.py"
SPEC = importlib.util.spec_from_file_location("schedule_night_shifts", SCRIPT)
assert SPEC and SPEC.loader
scheduler = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = scheduler
SPEC.loader.exec_module(scheduler)


def official_calendar(year: int) -> dict:
    return {
        "year": year,
        "holidays": [],
        "adjusted_workdays": [],
        "source_type": "official",
        "source_title": f"国务院办公厅关于{year}年部分节假日安排的通知",
        "source_url": "https://www.gov.cn/zhengce/test.htm",
        "publication_date": f"{year - 1}-11-01",
    }


def basic_request(year: int, month: int) -> dict:
    return {
        "year": year,
        "month": month,
        "staff": ["甲", "乙", "丙", "丁", "戊", "己", "庚"],
        "leaves": [],
        "no_night": [],
        "hard_constraints": [],
        "soft_preferences": [],
        "holiday_calendar": official_calendar(year),
    }


class RequestAndScoringTests(unittest.TestCase):
    def test_month_lengths_are_fully_covered_without_consecutive_nights(self):
        for year, month, expected_days in ((2026, 2, 28), (2028, 2, 29), (2026, 4, 30), (2026, 8, 31)):
            request = scheduler.normalize_request(basic_request(year, month))
            result = scheduler.solve_schedule(request, [], "recommended", 0.25)
            self.assertEqual(expected_days, len(result.schedule))
            self.assertEqual([], scheduler.validate_schedule(request, result.schedule))

    def test_2026_holiday_and_adjusted_weekend_scoring(self):
        calendar_data = official_calendar(2026)
        calendar_data["holidays"] = ["2026-01-01", "2026-01-02", "2026-01-03"]
        calendar_data["adjusted_workdays"] = ["2026-01-04"]
        self.assertEqual((4, "法定放假日"), scheduler.burden_for_date(scheduler.dt.date(2026, 1, 3), calendar_data))
        self.assertEqual((1, "调休上班日"), scheduler.burden_for_date(scheduler.dt.date(2026, 1, 4), calendar_data))
        self.assertEqual((0, "优质周四"), scheduler.burden_for_date(scheduler.dt.date(2026, 1, 8), calendar_data))

    def test_hard_conflict_is_rejected_with_date(self):
        raw = basic_request(2026, 8)
        raw["leaves"] = [{"name": "甲", "start": "2026-08-03", "end": "2026-08-03"}]
        raw["hard_constraints"] = [{"type": "required", "name": "甲", "dates": ["2026-08-03"]}]
        request = scheduler.normalize_request(raw)
        with self.assertRaisesRegex(scheduler.ScheduleError, "2026-08-03"):
            scheduler.solve_schedule(request, [], "recommended", 0.2)

    def test_excel_formula_like_staff_name_is_rejected(self):
        raw = basic_request(2026, 8)
        raw["staff"][0] = "=HYPERLINK(\"bad\")"
        with self.assertRaisesRegex(scheduler.ScheduleError, "Excel 不安全"):
            scheduler.normalize_request(raw)

    def test_previous_month_last_assignment_blocks_first_day(self):
        request = scheduler.normalize_request(basic_request(2026, 8))
        history = [
            {
                "month": "2026-07",
                "date": "2026-07-31",
                "staff": "甲",
                "burden": 1,
                "category": "普通工作日",
                "revision": 1,
                "source_url": "https://www.gov.cn/",
            }
        ]
        result = scheduler.solve_schedule(request, history, "recommended", 0.25)
        self.assertNotEqual("甲", result.schedule[0])
        self.assertEqual([], scheduler.validate_schedule(request, result.schedule, "甲"))


@unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl is required for workbook tests")
class WorkbookWorkflowTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.request_path = self.root / "request.json"
        self.candidates_path = self.root / "candidates.xlsx"
        self.confirmed_path = self.root / "confirmed.xlsx"
        self.revised_path = self.root / "revised.xlsx"

        request = basic_request(2026, 8)
        request["leaves"] = [{"name": "甲", "start": "2026-08-03", "end": "2026-08-05"}]
        request["hard_constraints"] = [{"type": "required", "name": "乙", "dates": ["2026-08-12"]}]
        request["soft_preferences"] = [{"type": "avoid_weekdays", "name": "丙", "weekdays": [6], "weight": 1}]
        self.request_path.write_text(json.dumps(request, ensure_ascii=False), encoding="utf-8")

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_generate_finalize_and_repair_auditable_workbook(self):
        import openpyxl

        scheduler.generate_workbook(self.request_path, self.candidates_path, None, 1.5)
        pending, pending_meta, pending_records = scheduler._load_workbook_state(self.candidates_path, "pending")
        self.assertEqual([], pending_records)
        self.assertIn("推荐方案", pending.sheetnames)
        self.assertIn("比较与校验", pending.sheetnames)
        self.assertIn("历史账本", pending.sheetnames)
        self.assertEqual("hidden", pending[scheduler.SYSTEM_SHEET].sheet_state)

        # 8月3日是甲请假，灰色应覆盖该工作日的默认底色。
        sheet = pending["推荐方案"]
        staff_column = 3
        leave_row = 5 + 2
        self.assertEqual("A6A6A6", sheet.cell(leave_row, staff_column).fill.fgColor.rgb[-6:])

        with self.assertRaisesRegex(scheduler.ScheduleError, "confirmed"):
            scheduler.generate_workbook(self.request_path, self.root / "invalid-history.xlsx", self.candidates_path, 0.5)

        scheduler.finalize_workbook(self.candidates_path, "recommended", self.confirmed_path)
        confirmed, confirmed_meta, records = scheduler._load_workbook_state(self.confirmed_path, "confirmed")
        self.assertIn("最终排班", confirmed.sheetnames)
        self.assertEqual("候选归档", confirmed["推荐方案"].cell(2, 2).value)
        self.assertEqual(31, len([record for record in records if record["month"] == "2026-08"]))
        original = json.loads(confirmed_meta["final_schedule_json"])

        target_index = next(index for index, name in enumerate(original) if name == "甲" and index > 5)
        target_date = f"2026-08-{target_index + 1:02d}"
        changes_path = self.root / "changes.json"
        changes_path.write_text(
            json.dumps({"leaves": [{"name": "甲", "start": target_date, "end": target_date}]}, ensure_ascii=False),
            encoding="utf-8",
        )
        scheduler.repair_workbook(self.confirmed_path, changes_path, self.revised_path, 1.5)
        revised, revised_meta, revised_records = scheduler._load_workbook_state(self.revised_path, "confirmed")
        repaired = json.loads(revised_meta["final_schedule_json"])
        self.assertNotEqual("甲", repaired[target_index])
        self.assertEqual("2", revised_meta["revision"])
        self.assertIn("改动记录", revised.sheetnames)
        self.assertEqual(31, len([record for record in revised_records if record["month"] == "2026-08"]))
        self.assertEqual([], scheduler.validate_schedule(json.loads(revised_meta["request_json"]), repaired))

        # Round-trip with openpyxl proves the generated file is a valid xlsx package.
        reopened = openpyxl.load_workbook(self.revised_path, data_only=False)
        self.assertIn("最终排班", reopened.sheetnames)

    def test_adjusted_sunday_stays_yellow_in_workbook(self):
        import openpyxl

        request = basic_request(2026, 1)
        request["holiday_calendar"]["holidays"] = ["2026-01-01", "2026-01-02", "2026-01-03"]
        request["holiday_calendar"]["adjusted_workdays"] = ["2026-01-04"]
        self.request_path.write_text(json.dumps(request, ensure_ascii=False), encoding="utf-8")
        scheduler.generate_workbook(self.request_path, self.candidates_path, None, 0.9)
        workbook = openpyxl.load_workbook(self.candidates_path)
        sheet = workbook["推荐方案"]
        january_four_row = 5 + 3
        self.assertEqual("FFF200", sheet.cell(january_four_row, 1).fill.fgColor.rgb[-6:])


if __name__ == "__main__":
    unittest.main()
