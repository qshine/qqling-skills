#!/usr/bin/env python3
"""Deterministic monthly night-shift scheduling and Excel audit workflow."""

from __future__ import annotations

import argparse
import calendar
import copy
import datetime as dt
import json
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Set, Tuple
from urllib.parse import urlparse


SCHEMA_VERSION = "1"
SYSTEM_SHEET = "_系统数据"
HISTORY_SHEET = "历史账本"
COMPARE_SHEET = "比较与校验"
FINAL_SHEET = "最终排班"

VARIANTS = {
    "recommended": "推荐方案",
    "fairness": "公平优先",
    "preference": "偏好优先",
    "repair": "修订方案",
}

CHOICE_ALIASES = {
    "recommended": "recommended",
    "a": "recommended",
    "推荐方案": "recommended",
    "fairness": "fairness",
    "b": "fairness",
    "公平优先": "fairness",
    "preference": "preference",
    "c": "preference",
    "偏好优先": "preference",
}

HARD_TYPES = {"required", "unavailable", "max_shifts", "min_shifts"}
SOFT_TYPES = {
    "avoid_dates",
    "prefer_dates",
    "avoid_weekdays",
    "prefer_weekdays",
    "max_shifts",
    "min_shifts",
}


class ScheduleError(ValueError):
    """A user-actionable scheduling or workbook validation error."""


@dataclass
class SearchResult:
    schedule: List[str]
    metrics: Dict[str, Any]


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:  # pragma: no cover - environment-specific
        raise ScheduleError(
            "缺少 openpyxl。请使用 Codex 工作区依赖加载器返回的 Python 运行时执行。"
        ) from exc
    return openpyxl


def _read_json(path: Path) -> Dict[str, Any]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ScheduleError(f"找不到 JSON 文件：{path}") from exc
    except json.JSONDecodeError as exc:
        raise ScheduleError(f"JSON 格式错误：{path}: {exc}") from exc
    if not isinstance(data, dict):
        raise ScheduleError("请求 JSON 顶层必须是对象。")
    return data


def _write_json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)


def _parse_date(value: Any, field: str) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    try:
        return dt.date.fromisoformat(str(value))
    except ValueError as exc:
        raise ScheduleError(f"{field} 必须使用 YYYY-MM-DD：{value}") from exc


def _month_key(date_value: dt.date) -> str:
    return f"{date_value.year:04d}-{date_value.month:02d}"


def _month_distance(target: dt.date, earlier: dt.date) -> int:
    return (target.year - earlier.year) * 12 + target.month - earlier.month


def _month_dates(year: int, month: int) -> List[dt.date]:
    return [dt.date(year, month, day) for day in range(1, calendar.monthrange(year, month)[1] + 1)]


def _ensure_month_date(value: Any, year: int, month: int, field: str) -> dt.date:
    parsed = _parse_date(value, field)
    if parsed.year != year or parsed.month != month:
        raise ScheduleError(f"{field} 必须位于 {year:04d}-{month:02d}：{parsed.isoformat()}")
    return parsed


def _normalize_staff(raw_staff: Any) -> Tuple[List[str], Set[str]]:
    if not isinstance(raw_staff, list) or not raw_staff:
        raise ScheduleError("staff 必须是非空数组。")
    names: List[str] = []
    inactive: Set[str] = set()
    for item in raw_staff:
        if isinstance(item, str):
            name = item.strip()
            active = True
        elif isinstance(item, dict):
            name = str(item.get("name", "")).strip()
            active = bool(item.get("active", True))
        else:
            raise ScheduleError("staff 项必须是姓名字符串或含 name 的对象。")
        if not name:
            raise ScheduleError("人员姓名不能为空。")
        if len(name) > 50 or name[0] in "=+-@" or any(ord(char) < 32 for char in name):
            raise ScheduleError(f"人员姓名包含 Excel 不安全字符或过长：{name!r}")
        if name in names:
            raise ScheduleError(f"人员名单包含重复姓名：{name}")
        names.append(name)
        if not active:
            inactive.add(name)
    return names, inactive


def _normalize_calendar(raw: Any, year: int) -> Dict[str, Any]:
    if not isinstance(raw, dict):
        raise ScheduleError("缺少 holiday_calendar；必须提供官方日历或用户确认的覆盖日期。")
    cal_year = int(raw.get("year", year))
    if cal_year != year:
        raise ScheduleError(f"holiday_calendar.year 必须为 {year}。")

    holidays = sorted({_parse_date(value, "holiday_calendar.holidays").isoformat() for value in raw.get("holidays", [])})
    workdays = sorted({_parse_date(value, "holiday_calendar.adjusted_workdays").isoformat() for value in raw.get("adjusted_workdays", [])})
    for value in holidays + workdays:
        if dt.date.fromisoformat(value).year != year:
            raise ScheduleError(f"节假日日历包含其他年份日期：{value}")
    if set(holidays) & set(workdays):
        overlap = sorted(set(holidays) & set(workdays))[0]
        raise ScheduleError(f"同一日期不能同时是放假日和调休上班日：{overlap}")

    source_type = str(raw.get("source_type", "official"))
    source_url = str(raw.get("source_url", "")).strip()
    confirmed = bool(raw.get("confirmed_by_user", False))
    if source_type == "official":
        host = (urlparse(source_url).hostname or "").lower()
        if host != "gov.cn" and not host.endswith(".gov.cn"):
            raise ScheduleError("官方节假日来源必须使用 gov.cn 域名。")
        if not raw.get("source_title"):
            raise ScheduleError("官方节假日来源缺少 source_title。")
    elif source_type == "user_override":
        if not confirmed:
            raise ScheduleError("用户覆盖日历必须设置 confirmed_by_user=true。")
    else:
        raise ScheduleError("holiday_calendar.source_type 仅支持 official 或 user_override。")

    return {
        "year": year,
        "holidays": holidays,
        "adjusted_workdays": workdays,
        "source_type": source_type,
        "source_title": str(raw.get("source_title", "用户确认的节假日日期")),
        "source_url": source_url,
        "publication_date": str(raw.get("publication_date", "")),
        "confirmed_by_user": confirmed,
    }


def normalize_request(raw: Mapping[str, Any], cached_calendar: Optional[Mapping[str, Any]] = None) -> Dict[str, Any]:
    try:
        year = int(raw["year"])
        month = int(raw["month"])
    except (KeyError, TypeError, ValueError) as exc:
        raise ScheduleError("请求必须包含有效的 year 和 month。") from exc
    if year < 2000 or year > 2200 or month < 1 or month > 12:
        raise ScheduleError("year 或 month 超出支持范围。")

    staff, inactive = _normalize_staff(raw.get("staff"))
    staff_set = set(staff)
    no_night = {str(name).strip() for name in raw.get("no_night", [])}
    unknown_no_night = no_night - staff_set
    if unknown_no_night:
        raise ScheduleError(f"no_night 包含未知人员：{', '.join(sorted(unknown_no_night))}")

    leaves: List[Dict[str, str]] = []
    for index, item in enumerate(raw.get("leaves", [])):
        if not isinstance(item, dict):
            raise ScheduleError(f"leaves[{index}] 必须是对象。")
        name = str(item.get("name", "")).strip()
        if name not in staff_set:
            raise ScheduleError(f"leaves[{index}] 包含未知人员：{name}")
        start = _ensure_month_date(item.get("start"), year, month, f"leaves[{index}].start")
        end = _ensure_month_date(item.get("end"), year, month, f"leaves[{index}].end")
        if end < start:
            raise ScheduleError(f"leaves[{index}] 结束日期早于开始日期。")
        leaves.append({"name": name, "start": start.isoformat(), "end": end.isoformat()})

    hard: List[Dict[str, Any]] = []
    for index, item in enumerate(raw.get("hard_constraints", [])):
        if not isinstance(item, dict):
            raise ScheduleError(f"hard_constraints[{index}] 必须是对象。")
        ctype = str(item.get("type", ""))
        if ctype not in HARD_TYPES:
            raise ScheduleError(f"不支持的硬规则类型：{ctype}")
        name = str(item.get("name", "")).strip()
        if name not in staff_set:
            raise ScheduleError(f"硬规则包含未知人员：{name}")
        normalized: Dict[str, Any] = {"type": ctype, "name": name}
        if ctype in {"required", "unavailable"}:
            values = item.get("dates", [])
            if not isinstance(values, list) or not values:
                raise ScheduleError(f"{ctype} 规则必须包含非空 dates。")
            normalized["dates"] = sorted(
                {_ensure_month_date(value, year, month, f"{ctype}.dates").isoformat() for value in values}
            )
        else:
            try:
                value = int(item.get("value"))
            except (TypeError, ValueError) as exc:
                raise ScheduleError(f"{ctype} 规则必须包含整数 value。") from exc
            if value < 0:
                raise ScheduleError(f"{ctype}.value 不能为负数。")
            normalized["value"] = value
        hard.append(normalized)

    soft: List[Dict[str, Any]] = []
    for index, item in enumerate(raw.get("soft_preferences", [])):
        if not isinstance(item, dict):
            raise ScheduleError(f"soft_preferences[{index}] 必须是对象。")
        ctype = str(item.get("type", ""))
        if ctype not in SOFT_TYPES:
            raise ScheduleError(f"不支持的软偏好类型：{ctype}")
        name = str(item.get("name", "")).strip()
        if name not in staff_set:
            raise ScheduleError(f"软偏好包含未知人员：{name}")
        try:
            weight = int(item.get("weight", 1))
        except (TypeError, ValueError) as exc:
            raise ScheduleError(f"soft_preferences[{index}].weight 必须是整数。") from exc
        if weight <= 0:
            raise ScheduleError("软偏好 weight 必须为正整数。")
        normalized = {"type": ctype, "name": name, "weight": weight}
        if ctype.endswith("_dates"):
            values = item.get("dates", [])
            if not isinstance(values, list) or not values:
                raise ScheduleError(f"{ctype} 必须包含非空 dates。")
            normalized["dates"] = sorted(
                {_ensure_month_date(value, year, month, f"{ctype}.dates").isoformat() for value in values}
            )
        elif ctype.endswith("_weekdays"):
            weekdays = sorted({int(value) for value in item.get("weekdays", [])})
            if not weekdays or any(value < 1 or value > 7 for value in weekdays):
                raise ScheduleError(f"{ctype}.weekdays 必须使用 1 至 7。")
            normalized["weekdays"] = weekdays
        else:
            try:
                value = int(item.get("value"))
            except (TypeError, ValueError) as exc:
                raise ScheduleError(f"{ctype} 必须包含整数 value。") from exc
            if value < 0:
                raise ScheduleError(f"{ctype}.value 不能为负数。")
            normalized["value"] = value
        soft.append(normalized)

    raw_calendar = raw.get("holiday_calendar")
    if raw_calendar is None and cached_calendar and int(cached_calendar.get("year", -1)) == year:
        raw_calendar = cached_calendar
    holiday_calendar = _normalize_calendar(raw_calendar, year)

    normalized_staff = [{"name": name, "active": name not in inactive} for name in staff]
    return {
        "year": year,
        "month": month,
        "staff": normalized_staff,
        "leaves": leaves,
        "no_night": sorted(no_night),
        "hard_constraints": hard,
        "soft_preferences": soft,
        "holiday_calendar": holiday_calendar,
    }


def merge_repair_request(base: Mapping[str, Any], changes: Mapping[str, Any]) -> Dict[str, Any]:
    if "replace_request" in changes:
        replacement = changes["replace_request"]
        if not isinstance(replacement, dict):
            raise ScheduleError("replace_request 必须是完整请求对象。")
        normalized = normalize_request(replacement, base.get("holiday_calendar"))
        base_names = [item["name"] for item in base["staff"]]
        new_names = [item["name"] for item in normalized["staff"]]
        if normalized["year"] != base["year"] or normalized["month"] != base["month"]:
            raise ScheduleError("修复不能改变已确认工作簿的年月。")
        if new_names != base_names:
            raise ScheduleError("修复不能改变人员顺序或名单；请重新生成该月排班。")
        return normalized

    merged = copy.deepcopy(dict(base))
    allowed = {"leaves", "no_night", "hard_constraints", "soft_preferences"}
    unknown = set(changes) - allowed
    if unknown:
        raise ScheduleError(f"修复请求包含未知字段：{', '.join(sorted(unknown))}")
    merged["leaves"] = list(merged.get("leaves", [])) + list(changes.get("leaves", []))
    merged["no_night"] = sorted(set(merged.get("no_night", [])) | set(changes.get("no_night", [])))
    merged["hard_constraints"] = list(merged.get("hard_constraints", [])) + list(changes.get("hard_constraints", []))
    merged["soft_preferences"] = list(merged.get("soft_preferences", [])) + list(changes.get("soft_preferences", []))
    return normalize_request(merged, base.get("holiday_calendar"))


def _staff_names(request: Mapping[str, Any]) -> List[str]:
    return [str(item["name"]) for item in request["staff"]]


def _active_staff(request: Mapping[str, Any]) -> List[str]:
    excluded = set(request.get("no_night", []))
    return [str(item["name"]) for item in request["staff"] if item.get("active", True) and item["name"] not in excluded]


def burden_for_date(date_value: dt.date, holiday_calendar: Mapping[str, Any]) -> Tuple[int, str]:
    value = date_value.isoformat()
    holidays = set(holiday_calendar.get("holidays", []))
    workdays = set(holiday_calendar.get("adjusted_workdays", []))
    if value in holidays:
        return 4, "法定放假日"
    if value in workdays:
        return 1, "调休上班日"
    if date_value.isoweekday() == 4:
        return 0, "优质周四"
    if date_value.isoweekday() == 6:
        return 3, "周六"
    if date_value.isoweekday() == 7:
        return 2, "周日"
    return 1, "普通工作日"


def _leave_dates(request: Mapping[str, Any]) -> Dict[str, Set[str]]:
    result = {name: set() for name in _staff_names(request)}
    for leave in request.get("leaves", []):
        start = dt.date.fromisoformat(leave["start"])
        end = dt.date.fromisoformat(leave["end"])
        current = start
        while current <= end:
            result[leave["name"]].add(current.isoformat())
            current += dt.timedelta(days=1)
    return result


def _constraint_maps(request: Mapping[str, Any]) -> Dict[str, Any]:
    names = _staff_names(request)
    days = len(_month_dates(request["year"], request["month"]))
    unavailable = _leave_dates(request)
    required: Dict[str, str] = {}
    maximum = {name: days for name in names}
    minimum = {name: 0 for name in names}
    for name in request.get("no_night", []):
        maximum[name] = 0
    for item in request.get("staff", []):
        if not item.get("active", True):
            maximum[item["name"]] = 0

    for rule in request.get("hard_constraints", []):
        name = rule["name"]
        if rule["type"] == "required":
            for value in rule["dates"]:
                if value in required and required[value] != name:
                    raise ScheduleError(f"{value} 同时要求 {required[value]} 和 {name} 值班。")
                required[value] = name
        elif rule["type"] == "unavailable":
            unavailable[name].update(rule["dates"])
        elif rule["type"] == "max_shifts":
            maximum[name] = min(maximum[name], int(rule["value"]))
        elif rule["type"] == "min_shifts":
            minimum[name] = max(minimum[name], int(rule["value"]))

    for value, name in required.items():
        if value in unavailable[name] or maximum[name] == 0:
            raise ScheduleError(f"{value} 要求 {name} 值班，但该人员当天不可排。")
    for name in names:
        if minimum[name] > maximum[name]:
            raise ScheduleError(f"{name} 的最少班次大于最多班次。")
    if sum(maximum.values()) < days:
        raise ScheduleError("所有人员的最多班次合计不足以覆盖整月。")
    if sum(minimum.values()) > days:
        raise ScheduleError("所有人员的最少班次合计超过当月天数。")

    dates = _month_dates(request["year"], request["month"])
    for left, right in zip(dates, dates[1:]):
        if required.get(left.isoformat()) and required.get(left.isoformat()) == required.get(right.isoformat()):
            raise ScheduleError(f"{left.isoformat()} 与 {right.isoformat()} 连续要求同一人值班。")

    return {
        "unavailable": unavailable,
        "required": required,
        "maximum": maximum,
        "minimum": minimum,
    }


def _history_loads(records: Sequence[Mapping[str, Any]], request: Mapping[str, Any]) -> Tuple[Dict[str, int], Dict[str, int]]:
    names = _staff_names(request)
    totals = {name: 0 for name in names}
    burdens = {name: 0 for name in names}
    target = dt.date(request["year"], request["month"], 1)
    weight_by_distance = {1: 3, 2: 2, 3: 1}
    for record in records:
        name = str(record.get("staff", ""))
        if name not in totals:
            continue
        record_date = _parse_date(record.get("date"), "历史日期")
        distance = _month_distance(target, record_date.replace(day=1))
        weight = weight_by_distance.get(distance)
        if weight:
            totals[name] += weight
            burdens[name] += weight * int(record.get("burden", 0))
    return totals, burdens


def _previous_month_last_staff(records: Sequence[Mapping[str, Any]], request: Mapping[str, Any]) -> Optional[str]:
    first = dt.date(request["year"], request["month"], 1)
    previous_last = first - dt.timedelta(days=1)
    for record in records:
        if _parse_date(record.get("date"), "历史日期") == previous_last:
            return str(record.get("staff"))
    return None


def _soft_increment(preferences: Sequence[Mapping[str, Any]], name: str, date_value: dt.date) -> int:
    cost = 0
    value = date_value.isoformat()
    weekday = date_value.isoweekday()
    for pref in preferences:
        weight = int(pref.get("weight", 1))
        if pref["type"] == "avoid_dates" and pref["name"] == name and value in pref["dates"]:
            cost += weight
        elif pref["type"] == "prefer_dates" and value in pref["dates"] and pref["name"] != name:
            cost += weight
        elif pref["type"] == "avoid_weekdays" and pref["name"] == name and weekday in pref["weekdays"]:
            cost += weight
        elif pref["type"] == "prefer_weekdays" and pref["name"] == name and weekday not in pref["weekdays"]:
            cost += weight
    return cost


def _soft_count_cost(preferences: Sequence[Mapping[str, Any]], counts: Mapping[str, int]) -> int:
    cost = 0
    for pref in preferences:
        weight = int(pref.get("weight", 1))
        if pref["type"] == "max_shifts":
            cost += max(0, counts[pref["name"]] - int(pref["value"])) * weight
        elif pref["type"] == "min_shifts":
            cost += max(0, int(pref["value"]) - counts[pref["name"]]) * weight
    return cost


def _unmet_preferences(request: Mapping[str, Any], schedule: Sequence[str]) -> List[str]:
    dates = _month_dates(request["year"], request["month"])
    counts = {name: schedule.count(name) for name in _staff_names(request)}
    unmet: List[str] = []
    by_date = dict(zip((value.isoformat() for value in dates), schedule))
    for pref in request.get("soft_preferences", []):
        name = pref["name"]
        ctype = pref["type"]
        if ctype == "avoid_dates":
            bad = [value for value in pref["dates"] if by_date[value] == name]
            if bad:
                unmet.append(f"{name} 未能避开日期：{', '.join(bad)}")
        elif ctype == "prefer_dates":
            bad = [value for value in pref["dates"] if by_date[value] != name]
            if bad:
                unmet.append(f"{name} 未排到希望日期：{', '.join(bad)}")
        elif ctype == "avoid_weekdays":
            bad = [date.isoformat() for date, assigned in zip(dates, schedule) if assigned == name and date.isoweekday() in pref["weekdays"]]
            if bad:
                unmet.append(f"{name} 未能避开星期偏好：{', '.join(bad)}")
        elif ctype == "prefer_weekdays":
            bad = [date.isoformat() for date, assigned in zip(dates, schedule) if assigned == name and date.isoweekday() not in pref["weekdays"]]
            if bad:
                unmet.append(f"{name} 有班次未落在希望星期：{', '.join(bad)}")
        elif ctype == "max_shifts" and counts[name] > int(pref["value"]):
            unmet.append(f"{name} 希望最多 {pref['value']} 次，实际 {counts[name]} 次")
        elif ctype == "min_shifts" and counts[name] < int(pref["value"]):
            unmet.append(f"{name} 希望至少 {pref['value']} 次，实际 {counts[name]} 次")
    return unmet


def validate_schedule(
    request: Mapping[str, Any],
    schedule: Sequence[str],
    previous_month_last: Optional[str] = None,
) -> List[str]:
    dates = _month_dates(request["year"], request["month"])
    names = _staff_names(request)
    maps = _constraint_maps(request)
    errors: List[str] = []
    if len(schedule) != len(dates):
        return [f"应有 {len(dates)} 天排班，实际 {len(schedule)} 天。"]
    counts = {name: 0 for name in names}
    previous = previous_month_last
    for date_value, name in zip(dates, schedule):
        value = date_value.isoformat()
        if name not in counts:
            errors.append(f"{value} 安排了未知人员：{name}")
            continue
        counts[name] += 1
        if value in maps["unavailable"][name]:
            errors.append(f"{value} 安排了不可值班人员：{name}")
        if maps["required"].get(value) not in {None, name}:
            errors.append(f"{value} 必须由 {maps['required'][value]} 值班。")
        if previous == name:
            errors.append(f"{value} 与前一天连续安排 {name}。")
        previous = name
    for name in names:
        if counts[name] > maps["maximum"][name]:
            errors.append(f"{name} 超过最多 {maps['maximum'][name]} 次。")
        if counts[name] < maps["minimum"][name]:
            errors.append(f"{name} 少于最少 {maps['minimum'][name]} 次。")
    return errors


def _evaluate(
    request: Mapping[str, Any],
    schedule: Sequence[str],
    records: Sequence[Mapping[str, Any]],
    previous_schedule: Optional[Sequence[str]] = None,
    variant: str = "recommended",
    nodes: int = 0,
    search_complete: bool = False,
) -> Dict[str, Any]:
    names = _staff_names(request)
    dates = _month_dates(request["year"], request["month"])
    counts = {name: 0 for name in names}
    burdens = {name: 0 for name in names}
    for date_value, name in zip(dates, schedule):
        counts[name] += 1
        burdens[name] += burden_for_date(date_value, request["holiday_calendar"])[0]
    historical_totals, historical_burdens = _history_loads(records, request)
    active = _active_staff(request)
    projected_totals = {name: historical_totals[name] + 3 * counts[name] for name in active}
    projected_burdens = {name: historical_burdens[name] + 3 * burdens[name] for name in active}
    total_spread = max(projected_totals.values()) - min(projected_totals.values()) if active else 0
    burden_spread = max(projected_burdens.values()) - min(projected_burdens.values()) if active else 0
    soft_penalty = sum(
        _soft_increment(request.get("soft_preferences", []), name, date_value)
        for name, date_value in zip(schedule, dates)
    ) + _soft_count_cost(request.get("soft_preferences", []), counts)
    changes = 0
    if previous_schedule is not None:
        changes = sum(left != right for left, right in zip(schedule, previous_schedule))
    return {
        "variant": variant,
        "counts": counts,
        "burdens": burdens,
        "projected_totals": projected_totals,
        "projected_burdens": projected_burdens,
        "total_spread": total_spread,
        "burden_spread": burden_spread,
        "soft_penalty": soft_penalty,
        "changes": changes,
        "unmet_preferences": _unmet_preferences(request, schedule),
        "nodes": nodes,
        "search_complete": search_complete,
    }


def _objective(metrics: Mapping[str, Any], schedule: Sequence[str], name_indexes: Mapping[str, int], variant: str) -> Tuple[Any, ...]:
    tie = tuple(name_indexes[name] for name in schedule)
    if variant == "fairness":
        return (metrics["burden_spread"], metrics["total_spread"], metrics["soft_penalty"], tie)
    if variant == "preference":
        return (metrics["soft_penalty"], metrics["total_spread"], metrics["burden_spread"], tie)
    if variant == "repair":
        return (metrics["changes"], metrics["total_spread"], metrics["burden_spread"], metrics["soft_penalty"], tie)
    return (metrics["total_spread"], metrics["burden_spread"], metrics["soft_penalty"], tie)


def _diagnose_no_solution(request: Mapping[str, Any], previous_month_last: Optional[str]) -> str:
    dates = _month_dates(request["year"], request["month"])
    active = _active_staff(request)
    maps = _constraint_maps(request)
    issues: List[str] = []
    for date_value in dates:
        value = date_value.isoformat()
        eligible = [name for name in active if value not in maps["unavailable"][name]]
        if maps["required"].get(value):
            eligible = [name for name in eligible if name == maps["required"][value]]
        if not eligible:
            issues.append(f"{value} 没有任何可值班人员")
    first_value = dates[0].isoformat()
    if previous_month_last and maps["required"].get(first_value) == previous_month_last:
        issues.append(f"{first_value} 要求 {previous_month_last} 值班，但其上月最后一天已值班")
    if not active:
        issues.append("没有可参与夜班的人员")
    detail = "；".join(issues[:5]) if issues else "规则组合、连续夜班限制或班次数上下限互相冲突"
    return f"找不到满足全部硬规则的排班：{detail}。"


def solve_schedule(
    request: Mapping[str, Any],
    records: Sequence[Mapping[str, Any]],
    variant: str,
    timeout_seconds: float,
    previous_schedule: Optional[Sequence[str]] = None,
) -> SearchResult:
    dates = _month_dates(request["year"], request["month"])
    names = _staff_names(request)
    active = _active_staff(request)
    if not active:
        raise ScheduleError("没有可参与夜班的人员。")
    maps = _constraint_maps(request)
    previous_month_last = _previous_month_last_staff(records, request)
    history_totals, history_burdens = _history_loads(records, request)
    burdens_by_day = [burden_for_date(value, request["holiday_calendar"])[0] for value in dates]
    indexes = {name: index for index, name in enumerate(names)}
    counts = {name: 0 for name in names}
    burdens = {name: 0 for name in names}
    schedule: List[str] = []
    best_schedule: Optional[List[str]] = None
    best_metrics: Optional[Dict[str, Any]] = None
    best_objective: Optional[Tuple[Any, ...]] = None
    deadline = time.monotonic() + max(0.05, timeout_seconds)
    nodes = 0
    timed_out = False
    dominance: Dict[Tuple[Any, ...], Tuple[int, int]] = {}
    partial_soft = 0
    partial_changes = 0

    def remaining_capacity(name: str, start_index: int) -> int:
        return sum(
            dates[index].isoformat() not in maps["unavailable"][name]
            for index in range(start_index, len(dates))
        )

    def forward_ok(next_index: int, last_name: Optional[str]) -> bool:
        for name in names:
            if counts[name] + remaining_capacity(name, next_index) < maps["minimum"][name]:
                return False
        if next_index < len(dates):
            value = dates[next_index].isoformat()
            required_name = maps["required"].get(value)
            if required_name:
                return (
                    required_name != last_name
                    and value not in maps["unavailable"][required_name]
                    and counts[required_name] < maps["maximum"][required_name]
                )
            return any(
                name != last_name
                and value not in maps["unavailable"][name]
                and counts[name] < maps["maximum"][name]
                for name in active
            )
        return True

    def candidate_order(index: int, last_name: Optional[str]) -> List[str]:
        value = dates[index].isoformat()
        required_name = maps["required"].get(value)
        candidates = [required_name] if required_name else list(active)
        eligible: List[str] = []
        for name in candidates:
            if name is None or name == last_name:
                continue
            if value in maps["unavailable"][name] or counts[name] >= maps["maximum"][name]:
                continue
            if index + 1 < len(dates) and maps["required"].get(dates[index + 1].isoformat()) == name:
                continue
            eligible.append(name)

        def key(name: str) -> Tuple[Any, ...]:
            projected_total = history_totals[name] + 3 * (counts[name] + 1)
            projected_burden = history_burdens[name] + 3 * (burdens[name] + burdens_by_day[index])
            local_soft = _soft_increment(request.get("soft_preferences", []), name, dates[index])
            local_change = int(previous_schedule is not None and previous_schedule[index] != name)
            if variant == "fairness":
                return projected_burden, projected_total, local_soft, indexes[name]
            if variant == "preference":
                return local_soft, projected_total, projected_burden, indexes[name]
            if variant == "repair":
                return local_change, projected_total, projected_burden, local_soft, indexes[name]
            return projected_total, projected_burden, local_soft, indexes[name]

        return sorted(eligible, key=key)

    def search(index: int, last_name: Optional[str]) -> None:
        nonlocal nodes, timed_out, best_schedule, best_metrics, best_objective
        nonlocal partial_soft, partial_changes
        nodes += 1
        if nodes % 256 == 0 and time.monotonic() >= deadline:
            timed_out = True
            return
        if timed_out:
            return
        if index == len(dates):
            if any(counts[name] < maps["minimum"][name] for name in names):
                return
            metrics = _evaluate(request, schedule, records, previous_schedule, variant, nodes, False)
            score = _objective(metrics, schedule, indexes, variant)
            if best_objective is None or score < best_objective:
                best_objective = score
                best_schedule = list(schedule)
                best_metrics = metrics
            return

        state = (
            index,
            last_name,
            tuple(counts[name] for name in names),
            tuple(burdens[name] for name in names),
        )
        prior = dominance.get(state)
        if prior is not None and prior[0] <= partial_soft and prior[1] <= partial_changes:
            return
        dominance[state] = (partial_soft, partial_changes)

        if best_objective is not None:
            if variant == "preference" and partial_soft > best_objective[0]:
                return
            if variant == "repair" and partial_changes > best_objective[0]:
                return

        candidates = candidate_order(index, last_name)
        for name in candidates:
            date_value = dates[index]
            soft_delta = _soft_increment(request.get("soft_preferences", []), name, date_value)
            change_delta = int(previous_schedule is not None and previous_schedule[index] != name)
            schedule.append(name)
            counts[name] += 1
            burdens[name] += burdens_by_day[index]
            partial_soft += soft_delta
            partial_changes += change_delta
            if forward_ok(index + 1, name):
                search(index + 1, name)
            partial_changes -= change_delta
            partial_soft -= soft_delta
            burdens[name] -= burdens_by_day[index]
            counts[name] -= 1
            schedule.pop()
            if timed_out:
                break

    search(0, previous_month_last)
    if best_schedule is None or best_metrics is None:
        raise ScheduleError(_diagnose_no_solution(request, previous_month_last))
    errors = validate_schedule(request, best_schedule, previous_month_last)
    if errors:
        raise ScheduleError("内部校验失败：" + "；".join(errors))
    best_metrics = _evaluate(
        request,
        best_schedule,
        records,
        previous_schedule,
        variant,
        nodes,
        not timed_out,
    )
    return SearchResult(best_schedule, best_metrics)


def _load_system_metadata(workbook: Any) -> Dict[str, str]:
    if SYSTEM_SHEET not in workbook.sheetnames:
        raise ScheduleError("工作簿不是 schedule-night-shifts 生成的文件：缺少系统数据。")
    sheet = workbook[SYSTEM_SHEET]
    metadata: Dict[str, str] = {}
    for key, value in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if key is not None:
            metadata[str(key)] = "" if value is None else str(value)
    if metadata.get("schema_version") != SCHEMA_VERSION:
        raise ScheduleError("工作簿数据版本不受支持。")
    return metadata


def _write_system_metadata(workbook: Any, metadata: Mapping[str, Any]) -> None:
    if SYSTEM_SHEET in workbook.sheetnames:
        del workbook[SYSTEM_SHEET]
    sheet = workbook.create_sheet(SYSTEM_SHEET)
    for row, key in enumerate(sorted(metadata), start=1):
        value = metadata[key]
        sheet.cell(row, 1, key)
        sheet.cell(row, 2, value if isinstance(value, str) else _write_json_text(value))
    sheet.sheet_state = "hidden"


def _read_history_records(workbook: Any) -> List[Dict[str, Any]]:
    if HISTORY_SHEET not in workbook.sheetnames:
        return []
    sheet = workbook[HISTORY_SHEET]
    headers = [sheet.cell(4, column).value for column in range(1, 8)]
    expected = ["月份", "日期", "人员", "负担分", "日期类别", "修订版", "节假日来源"]
    if headers != expected:
        raise ScheduleError("历史账本表头不受支持或已被修改。")
    records: List[Dict[str, Any]] = []
    source_by_month: Dict[str, str] = {}
    for row in sheet.iter_rows(min_row=5, max_col=7, values_only=True):
        if row[1] in {None, ""}:
            continue
        date_value = _parse_date(row[1], "历史账本日期")
        month = str(row[0] or _month_key(date_value))
        source_value = str(row[6] or "")
        if source_value == "同上":
            source_value = source_by_month.get(month, "")
        elif source_value:
            source_by_month[month] = source_value
        records.append(
            {
                "month": month,
                "date": date_value.isoformat(),
                "staff": str(row[2]),
                "burden": int(row[3]),
                "category": str(row[4]),
                "revision": int(row[5]),
                "source_url": source_value,
            }
        )
    return records


def _load_workbook_state(path: Path, required_status: Optional[str] = None) -> Tuple[Any, Dict[str, str], List[Dict[str, Any]]]:
    openpyxl = _require_openpyxl()
    if not path.exists():
        raise ScheduleError(f"找不到工作簿：{path}")
    workbook = openpyxl.load_workbook(path)
    metadata = _load_system_metadata(workbook)
    if required_status and metadata.get("status") != required_status:
        raise ScheduleError(f"工作簿状态必须为 {required_status}，当前为 {metadata.get('status', 'unknown')}。")
    return workbook, metadata, _read_history_records(workbook)


def _cell_styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore

    thin = Side(style="thin", color="A6A6A6")
    medium = Side(style="medium", color="666666")
    return {
        "title_fill": PatternFill("solid", fgColor="1F4E78"),
        "header_fill": PatternFill("solid", fgColor="D9EAF7"),
        "weekend_fill": PatternFill("solid", fgColor="FFF200"),
        "leave_fill": PatternFill("solid", fgColor="A6A6A6"),
        "good_fill": PatternFill("solid", fgColor="E2F0D9"),
        "warn_fill": PatternFill("solid", fgColor="FCE4D6"),
        "white_font": Font(name="Microsoft YaHei", color="FFFFFF", bold=True, size=15),
        "header_font": Font(name="Microsoft YaHei", bold=True, size=10),
        "body_font": Font(name="Microsoft YaHei", size=10),
        "dot_font": Font(name="Arial", bold=True, size=15, color="000000"),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "grid": Border(left=thin, right=thin, top=thin, bottom=thin),
        "section": Border(bottom=medium),
    }


def _weekday_cn(value: dt.date) -> str:
    return "一二三四五六日"[value.isoweekday() - 1]


def _render_schedule_sheet(
    workbook: Any,
    sheet_name: str,
    request: Mapping[str, Any],
    schedule: Sequence[str],
    metrics: Mapping[str, Any],
    status: str,
    index: Optional[int] = None,
) -> Any:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name, index if index is not None else len(workbook.sheetnames))
    styles = _cell_styles()
    names = _staff_names(request)
    dates = _month_dates(request["year"], request["month"])
    leaves = _leave_dates(request)
    last_column = 2 + len(names)
    title = f"{request['year']}年{request['month']}月科室夜班表 — {sheet_name}"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = sheet.cell(1, 1, title)
    title_cell.fill = styles["title_fill"]
    title_cell.font = styles["white_font"]
    title_cell.alignment = styles["center"]
    sheet.row_dimensions[1].height = 28

    sheet.cell(2, 1, "状态")
    status_labels = {"pending": "待确认", "confirmed": "已确认", "archived": "候选归档"}
    sheet.cell(2, 2, status_labels.get(status, status))
    sheet.cell(2, 3, f"总次数差：{metrics['total_spread']}")
    sheet.cell(2, 4, f"负担差：{metrics['burden_spread']}")
    sheet.cell(2, 5, f"未满足偏好：{len(metrics['unmet_preferences'])}")
    sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)
    sheet.cell(3, 1, "图例：● 夜班｜黄色 周末或法定放假日｜灰色 个人请假（灰色优先）")
    sheet.cell(3, 1).alignment = styles["left"]

    headers = ["日期", "星期"] + names
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(4, column, header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["grid"]

    holidays = set(request["holiday_calendar"]["holidays"])
    for offset, (date_value, assigned) in enumerate(zip(dates, schedule), start=5):
        value = date_value.isoformat()
        is_yellow = date_value.isoweekday() in {6, 7} or value in holidays
        sheet.cell(offset, 1, date_value)
        sheet.cell(offset, 1).number_format = "yyyy-mm-dd"
        sheet.cell(offset, 2, _weekday_cn(date_value))
        for column in range(1, last_column + 1):
            cell = sheet.cell(offset, column)
            cell.font = styles["body_font"]
            cell.alignment = styles["center"]
            cell.border = styles["grid"]
            if is_yellow:
                cell.fill = styles["weekend_fill"]
        for staff_index, name in enumerate(names, start=3):
            cell = sheet.cell(offset, staff_index)
            if value in leaves[name]:
                cell.fill = styles["leave_fill"]
            if assigned == name:
                cell.value = "●"
                cell.font = styles["dot_font"]
        sheet.row_dimensions[offset].height = 22

    sheet.column_dimensions["A"].width = 13
    sheet.column_dimensions["B"].width = 7
    for column in range(3, last_column + 1):
        sheet.column_dimensions[sheet.cell(4, column).column_letter].width = 12
    sheet.freeze_panes = "C5"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:4"
    sheet.print_area = f"A1:{sheet.cell(4 + len(dates), last_column).coordinate}"
    return sheet


def _render_compare_sheet(workbook: Any, request: Mapping[str, Any], candidates: Mapping[str, Sequence[str]], metrics: Mapping[str, Mapping[str, Any]], status: str, selected: str = "") -> Any:
    if COMPARE_SHEET in workbook.sheetnames:
        del workbook[COMPARE_SHEET]
    sheet = workbook.create_sheet(COMPARE_SHEET)
    styles = _cell_styles()
    sheet.merge_cells("A1:I1")
    sheet["A1"] = "候选方案比较与校验"
    sheet["A1"].fill = styles["title_fill"]
    sheet["A1"].font = styles["white_font"]
    sheet["A1"].alignment = styles["center"]
    sheet["A2"] = "工作簿状态"
    sheet["B2"] = "已确认" if status == "confirmed" else "待确认"
    sheet["C2"] = "已选方案"
    sheet["D2"] = selected or "—"
    headers = ["方案", "总次数差", "负担差", "偏好罚分", "改动天数", "硬规则", "搜索节点", "搜索完成", "未满足偏好"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(4, column, header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["grid"]
    for row, key in enumerate(candidates, start=5):
        item = metrics[key]
        errors = validate_schedule(request, candidates[key])
        values = [
            VARIANTS.get(key, key),
            item["total_spread"],
            item["burden_spread"],
            item["soft_penalty"],
            item.get("changes", 0),
            "通过" if not errors else "失败",
            item.get("nodes", 0),
            "是" if item.get("search_complete") else "否（已返回最佳可行解）",
            "；".join(item.get("unmet_preferences", [])) or "无",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.border = styles["grid"]
            cell.alignment = styles["left"] if column == 9 else styles["center"]
    source = request["holiday_calendar"]
    source_row = 7 + len(candidates)
    sheet.cell(source_row, 1, "节假日来源")
    sheet.cell(source_row, 2, source.get("source_title", ""))
    sheet.cell(source_row + 1, 1, "来源网址")
    sheet.cell(source_row + 1, 2, source.get("source_url", "用户确认覆盖"))
    sheet.cell(source_row + 2, 1, "发布日期")
    sheet.cell(source_row + 2, 2, source.get("publication_date", ""))
    sheet.cell(source_row + 4, 1, "公平规则")
    sheet.cell(source_row + 4, 2, "近三个月权重 3:2:1；当前月按权重 3；周四0、普通工作日1、周日2、周六3、法定放假日4。")
    sheet.cell(source_row + 5, 1, "确认要求")
    sheet.cell(source_row + 5, 2, "候选方案不会写入历史。必须运行 finalize 明确确认后才能作为下月历史。")
    for row in range(1, source_row + 6):
        sheet.cell(row, 1).font = styles["header_font"] if row in {1, 2, 4, source_row, source_row + 1, source_row + 2, source_row + 4, source_row + 5} else styles["body_font"]
    widths = [16, 18, 14, 14, 12, 12, 12, 20, 56]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(4, column).column_letter].width = width
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    return sheet


def _render_history_sheet(workbook: Any, records: Sequence[Mapping[str, Any]], request: Mapping[str, Any]) -> Any:
    if HISTORY_SHEET in workbook.sheetnames:
        del workbook[HISTORY_SHEET]
    sheet = workbook.create_sheet(HISTORY_SHEET)
    styles = _cell_styles()
    sheet.merge_cells("A1:G1")
    sheet["A1"] = "已确认夜班历史账本"
    sheet["A1"].fill = styles["title_fill"]
    sheet["A1"].font = styles["white_font"]
    sheet["A1"].alignment = styles["center"]
    sheet.merge_cells("A2:G2")
    sheet["A2"] = "仅 finalize 或 repair 的最终方案进入账本；请勿手工改动表头。"
    headers = ["月份", "日期", "人员", "负担分", "日期类别", "修订版", "节假日来源"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(4, column, header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = styles["center"]
        cell.border = styles["grid"]
    sorted_records = sorted(records, key=lambda item: (str(item["date"]), int(item.get("revision", 1))))
    rendered_source_months: Set[str] = set()
    for row, record in enumerate(sorted_records, start=5):
        month = str(record["month"])
        source_display = record.get("source_url", "") if month not in rendered_source_months else "同上"
        rendered_source_months.add(month)
        values = [
            month,
            dt.date.fromisoformat(str(record["date"])),
            record["staff"],
            int(record["burden"]),
            record["category"],
            int(record["revision"]),
            source_display,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.border = styles["grid"]
            cell.alignment = styles["center"] if column != 7 else styles["left"]
        sheet.cell(row, 2).number_format = "yyyy-mm-dd"

    recent_months = sorted({str(record["month"]) for record in records}, reverse=True)[:3]
    recent = [record for record in records if str(record["month"]) in recent_months]
    sheet["I1"] = "近三个月摘要"
    sheet["I1"].fill = styles["title_fill"]
    sheet["I1"].font = styles["white_font"]
    for column, header in enumerate(["人员", "夜班数", "负担分"], start=9):
        cell = sheet.cell(4, column, header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["grid"]
    for row, name in enumerate(_staff_names(request), start=5):
        person_records = [item for item in recent if item["staff"] == name]
        values = [name, len(person_records), sum(int(item["burden"]) for item in person_records)]
        for column, value in enumerate(values, start=9):
            cell = sheet.cell(row, column, value)
            cell.border = styles["grid"]
            cell.alignment = styles["center"]
    widths = {"A": 11, "B": 13, "C": 14, "D": 10, "E": 16, "F": 10, "G": 48, "I": 14, "J": 12, "K": 12}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    return sheet


def _schedule_records(request: Mapping[str, Any], schedule: Sequence[str], revision: int) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for date_value, name in zip(_month_dates(request["year"], request["month"]), schedule):
        burden, category = burden_for_date(date_value, request["holiday_calendar"])
        records.append(
            {
                "month": _month_key(date_value),
                "date": date_value.isoformat(),
                "staff": name,
                "burden": burden,
                "category": category,
                "revision": revision,
                "source_url": request["holiday_calendar"].get("source_url", ""),
            }
        )
    return records


def _ensure_output_path(input_paths: Iterable[Path], output: Path) -> None:
    resolved_output = output.expanduser().resolve()
    for value in input_paths:
        if value.expanduser().resolve() == resolved_output:
            raise ScheduleError("输出路径不能覆盖输入文件；请使用新的文件名。")
    output.parent.mkdir(parents=True, exist_ok=True)


def generate_workbook(request_path: Path, output: Path, history_path: Optional[Path], timeout_seconds: float) -> None:
    history_records: List[Dict[str, Any]] = []
    cached_calendar: Optional[Dict[str, Any]] = None
    if history_path:
        _, history_meta, history_records = _load_workbook_state(history_path, "confirmed")
        previous_request = json.loads(history_meta["request_json"])
        cached_calendar = previous_request.get("holiday_calendar")
    request = normalize_request(_read_json(request_path), cached_calendar)
    target_month = f"{request['year']:04d}-{request['month']:02d}"
    if any(record["month"] == target_month for record in history_records):
        raise ScheduleError(f"历史账本已包含 {target_month}；请使用 repair，而不是重复 generate。")

    timeout_each = max(0.1, timeout_seconds / 3.0)
    candidates: Dict[str, List[str]] = {}
    metrics: Dict[str, Dict[str, Any]] = {}
    seen: Set[Tuple[str, ...]] = set()
    for variant in ("recommended", "fairness", "preference"):
        result = solve_schedule(request, history_records, variant, timeout_each)
        key = tuple(result.schedule)
        if key in seen:
            continue
        seen.add(key)
        candidates[variant] = result.schedule
        metrics[variant] = result.metrics
    if not candidates:
        raise ScheduleError("未生成任何可行候选方案。")

    openpyxl = _require_openpyxl()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for variant, schedule in candidates.items():
        _render_schedule_sheet(workbook, VARIANTS[variant], request, schedule, metrics[variant], "pending")
    _render_compare_sheet(workbook, request, candidates, metrics, "pending")
    _render_history_sheet(workbook, history_records, request)
    _write_system_metadata(
        workbook,
        {
            "schema_version": SCHEMA_VERSION,
            "status": "pending",
            "request_json": _write_json_text(request),
            "candidates_json": _write_json_text(candidates),
            "metrics_json": _write_json_text(metrics),
            "selected": "",
            "revision": "0",
        },
    )
    _ensure_output_path([request_path] + ([history_path] if history_path else []), output)
    workbook.save(output)


def finalize_workbook(workbook_path: Path, choice: str, output: Path) -> None:
    workbook, metadata, records = _load_workbook_state(workbook_path, "pending")
    try:
        request = json.loads(metadata["request_json"])
        candidates: Dict[str, List[str]] = json.loads(metadata["candidates_json"])
        metrics: Dict[str, Dict[str, Any]] = json.loads(metadata["metrics_json"])
    except (KeyError, json.JSONDecodeError) as exc:
        raise ScheduleError("候选工作簿的系统数据损坏。") from exc
    selected = CHOICE_ALIASES.get(choice.strip().lower(), CHOICE_ALIASES.get(choice.strip()))
    if selected not in candidates:
        available = ", ".join(VARIANTS.get(key, key) for key in candidates)
        raise ScheduleError(f"所选方案不存在。可选方案：{available}")
    schedule = candidates[selected]
    previous_last = _previous_month_last_staff(records, request)
    errors = validate_schedule(request, schedule, previous_last)
    if errors:
        raise ScheduleError("确认前校验失败：" + "；".join(errors))

    _render_schedule_sheet(workbook, FINAL_SHEET, request, schedule, metrics[selected], "confirmed", 0)
    for variant in candidates:
        sheet_name = VARIANTS.get(variant, variant)
        if sheet_name in workbook.sheetnames:
            workbook[sheet_name].cell(2, 2, "候选归档")
    records = [record for record in records if record["month"] != f"{request['year']:04d}-{request['month']:02d}"]
    records.extend(_schedule_records(request, schedule, 1))
    _render_compare_sheet(workbook, request, candidates, metrics, "confirmed", VARIANTS[selected])
    _render_history_sheet(workbook, records, request)
    _write_system_metadata(
        workbook,
        {
            "schema_version": SCHEMA_VERSION,
            "status": "confirmed",
            "request_json": _write_json_text(request),
            "candidates_json": _write_json_text(candidates),
            "metrics_json": _write_json_text(metrics),
            "selected": selected,
            "final_schedule_json": _write_json_text(schedule),
            "revision": "1",
        },
    )
    _ensure_output_path([workbook_path], output)
    workbook.save(output)


def _render_changes_sheet(workbook: Any, request: Mapping[str, Any], old: Sequence[str], new: Sequence[str], revision: int) -> None:
    name = "改动记录"
    if name in workbook.sheetnames:
        del workbook[name]
    sheet = workbook.create_sheet(name, 1)
    styles = _cell_styles()
    sheet.merge_cells("A1:D1")
    sheet["A1"] = f"排班修订记录 v{revision}"
    sheet["A1"].fill = styles["title_fill"]
    sheet["A1"].font = styles["white_font"]
    sheet["A1"].alignment = styles["center"]
    headers = ["日期", "原安排", "新安排", "说明"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.border = styles["grid"]
        cell.alignment = styles["center"]
    unavailable = _constraint_maps(request)["unavailable"]
    row = 4
    for date_value, before, after in zip(_month_dates(request["year"], request["month"]), old, new):
        if before == after:
            continue
        reason = "原安排触发新增硬约束" if date_value.isoformat() in unavailable[before] else "为保持全部硬规则且尽量减少改动"
        values = [date_value, before, after, reason]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.border = styles["grid"]
            cell.alignment = styles["center"] if column != 4 else styles["left"]
        sheet.cell(row, 1).number_format = "yyyy-mm-dd"
        row += 1
    if row == 4:
        sheet.cell(4, 1, "无改动")
        sheet.merge_cells("A4:D4")
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 40
    sheet.sheet_view.showGridLines = False


def repair_workbook(workbook_path: Path, changes_path: Path, output: Path, timeout_seconds: float) -> None:
    workbook, metadata, records = _load_workbook_state(workbook_path, "confirmed")
    try:
        base_request = json.loads(metadata["request_json"])
        old_schedule: List[str] = json.loads(metadata["final_schedule_json"])
        revision = int(metadata.get("revision", "1")) + 1
    except (KeyError, ValueError, json.JSONDecodeError) as exc:
        raise ScheduleError("确认工作簿的系统数据损坏。") from exc
    request = merge_repair_request(base_request, _read_json(changes_path))
    target_month = f"{request['year']:04d}-{request['month']:02d}"
    prior_records = [record for record in records if record["month"] != target_month]
    result = solve_schedule(request, prior_records, "repair", timeout_seconds, old_schedule)
    previous_last = _previous_month_last_staff(prior_records, request)
    errors = validate_schedule(request, result.schedule, previous_last)
    if errors:
        raise ScheduleError("修复后校验失败：" + "；".join(errors))

    _render_schedule_sheet(workbook, FINAL_SHEET, request, result.schedule, result.metrics, "confirmed", 0)
    _render_changes_sheet(workbook, request, old_schedule, result.schedule, revision)
    records = prior_records + _schedule_records(request, result.schedule, revision)
    _render_history_sheet(workbook, records, request)
    repair_candidates = {"repair": result.schedule}
    repair_metrics = {"repair": result.metrics}
    _render_compare_sheet(workbook, request, repair_candidates, repair_metrics, "confirmed", f"修订版 v{revision}")
    _write_system_metadata(
        workbook,
        {
            "schema_version": SCHEMA_VERSION,
            "status": "confirmed",
            "request_json": _write_json_text(request),
            "candidates_json": metadata.get("candidates_json", "{}"),
            "metrics_json": metadata.get("metrics_json", "{}"),
            "selected": metadata.get("selected", ""),
            "final_schedule_json": _write_json_text(result.schedule),
            "revision": str(revision),
        },
    )
    _ensure_output_path([workbook_path, changes_path], output)
    workbook.save(output)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="生成、确认和修复科室夜班 Excel 排班表。")
    subparsers = parser.add_subparsers(dest="command", required=True)

    generate = subparsers.add_parser("generate", help="生成候选排班工作簿")
    generate.add_argument("--request", type=Path, required=True)
    generate.add_argument("--output", type=Path, required=True)
    generate.add_argument("--history", type=Path)
    generate.add_argument("--timeout", type=float, default=30.0)

    finalize = subparsers.add_parser("finalize", help="确认候选方案并写入历史")
    finalize.add_argument("--workbook", type=Path, required=True)
    finalize.add_argument("--choice", required=True)
    finalize.add_argument("--output", type=Path, required=True)

    repair = subparsers.add_parser("repair", help="最小改动修复已确认排班")
    repair.add_argument("--workbook", type=Path, required=True)
    repair.add_argument("--changes", type=Path, required=True)
    repair.add_argument("--output", type=Path, required=True)
    repair.add_argument("--timeout", type=float, default=30.0)
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.command == "generate":
            generate_workbook(args.request, args.output, args.history, args.timeout)
        elif args.command == "finalize":
            finalize_workbook(args.workbook, args.choice, args.output)
        elif args.command == "repair":
            repair_workbook(args.workbook, args.changes, args.output, args.timeout)
        else:  # pragma: no cover
            raise ScheduleError(f"未知命令：{args.command}")
    except ScheduleError as exc:
        print(f"错误：{exc}", file=sys.stderr)
        return 2
    print(f"已生成：{args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
